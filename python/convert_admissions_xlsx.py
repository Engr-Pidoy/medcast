"""
Convert hospital admissions Excel (.xlsx) to MEDCAST CSV.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[wb.sheetnames[0]]

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)

    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "date",
                "day",
                "admissions",
                "discharges",
                "occupied_beds",
                "bed_capacity",
                "occupancy_rate",
            ]
        )

        # Detect header row
        rows = list(ws.iter_rows(values_only=True))
        start = 0
        if rows and rows[0] and str(rows[0][0]).lower().startswith("date"):
            start = 1

        count = 0
        for row in rows[start:]:
            if not row or row[0] is None or row[2] is None:
                continue
            d = row[0]
            date = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
            occ = row[6]
            writer.writerow(
                [
                    date,
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    round(float(occ), 2) if occ is not None else None,
                ]
            )
            count += 1

    print(f"converted={count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
